# -*- coding: utf-8 -*-
"""
경영지원부 주간 업무 보고서 생성기
-----------------------------------
사용법:
  1) '주간보고서_만들기.bat' 더블클릭 (또는 이 파일을 파이썬으로 실행)
  2) 창이 뜨면 NOWSYSTEM에서 내보낸 주간보고 엑셀(.xls) 3개를 선택
  3) 보고일(월요일) 확인
  4) 같은 폴더에 '경영지원부_주간보고_YYYYMMDD.xlsx' 가 자동 생성/열림

특징:
  - 입력 3개 엑셀의 모든 업무 내용을 누락 없이 100% 반영
  - PDF 보고서와 동일한 양식(머리글/체크리스트/구분별 본문)으로 출력
  - 담당자 표시 없이 '구분(카테고리)'별로 3명 업무를 병합
"""

import os
import re
import io
import sys
import json
import html
import difflib
import datetime
import traceback
import urllib.request
import urllib.error

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

try:
    import windnd            # 드래그&드롭 (Windows)
except Exception:
    windnd = None

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    tk.Tk().withdraw()
    messagebox.showerror("오류", "openpyxl 라이브러리가 필요합니다.\n명령창에서  pip install openpyxl  실행 후 다시 시도하세요.")
    sys.exit(1)

try:    # 이미지 삽입용 (Pillow 필요). 없으면 이미지 없이 동작.
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    XLImage = None


# ===================================================================
#  [설정] 매주 고정되는 부분 (필요시 이 부분만 수정하면 됩니다)
# ===================================================================
부서명 = "경영지원부"
보고자 = "이동숙 부장"

# 상단 '업무 체크리스트' 표 (고정) : (업무내용, 주기, 담당자)
체크리스트 = [
    ("팩스함 확인 및 서버 저장\n[위치 : \\\\192.168.0.221\\경영지원_부서\\팩스수신함]", "일 2회", "박선하"),
    ("매입매출 계산서 정리 및 그룹웨어 업로드 (업로드 후 공지)\n[위치 : 그룹웨어 – 계약관리 - 매입매출]", "주 1회", "박선하"),
    ("보증보험리스트 업데이트 및 그룹웨어 업로드\n[위치 : 그룹웨어 – 계약관리 – 보증보험 리스트]", "주 1회", "박선하"),
    ("입찰리스트 업데이트 및 업로드\n[위치 : 그룹웨어 – 영업관리 – 입찰]", "일 1회", "김유미"),
]

# 본문 첫 줄 '데일리 업무' (고정 템플릿)
데일리_지난주 = ("- 팩스 확인 및 서버 저장\n : 매일 오전,오후 확인\n"
              "- 우편물 수거 및 정리\n : 매주 수,금 확인\n"
              "- 우편 발송 및 택배 발송 \n- 더존 작업\n"
              "- 보증보험 리스트 작성 및 업데이트\n"
              "- 세금계산서 업로드\n  : 매주 금 업로드")
데일리_진행률 = "100%"
데일리_금주 = ("- 팩스 확인 및 서버 저장\n- 우편물 수거 및 정리\n"
            "- 우편 발송 및 택배 발송 \n- 더존 작업\n"
            "- 보증보험 리스트 작성 및 업데이트\n- 세금계산서 업로드")

# '데일리 업무'(상단 고정)에 이미 들어있는 항목들.
#  본문 아래에 이와 같은(중복) 업무가 있으면 팝업으로 '뺄지' 확인합니다.
데일리_항목 = [
    "팩스 확인 및 서버 저장",
    "우편물 수거",
    "우편 발송 및 택배 발송",
    "더존 작업",
    "보증보험 리스트 작성 및 업데이트",
    "세금계산서 업로드",
]

# 입력의 '업무분류' -> 출력 '구분(카테고리)' 매핑 및 표시 순서
구분_매핑 = {
    "경영": "경영관리",
    "재무": "재무업무",
    "입찰": "조달업무/입찰업무",
    "조달": "조달업무/입찰업무",
    "지원": "지원사업",
    "R&D": "R&D사업",
    "현장": "현장",
    "기타": "기타업무",
    "요청사항": "요청사항",
    "신규": "신규",
}
구분_순서 = ["경영관리", "재무업무", "조달업무/입찰업무", "지원사업",
          "R&D사업", "현장", "기타업무", "요청사항", "신규"]

# 게시판 항목 중 '재무업무 맨 위에 고정'으로 올릴 항목(라벨에 포함된 단어로 판별)
#  - 이름 제거 후 같은 내용은 한 줄로 합쳐 중복 표시되지 않게 함
재무_고정_게시판 = ["세금계산서", "분납서"]
재무_고정_구분 = "재무업무"

# 보고서 본문/게시판에서 제거할 사람 이름(게시판에 누가 작성했는지 표시되는 [이름] 태그)
#  - '업무 보고자' 명의는 그대로 유지됩니다.
이름목록 = ["이동숙", "박선하", "김유미"]

# '조달업무/입찰업무 맨 위에 고정'으로 올릴 표준 업무 (입력에서 키워드로 찾아 위로 올림)
#  - 공백 무시하고 부분일치. 목록 순서대로 맨 위에 배치.
#  - 매칭이 안 되면(텍스트가 달라지면) 이 키워드만 실제 문구에 맞게 고치면 됩니다.
조달입찰_고정_구분 = "조달업무/입찰업무"
조달입찰_고정_키워드 = [
    "검색리스트",            # 입찰 및 입찰 결과 업데이트 … 입찰 검색리스트 작성
    "입찰업무정리",          # 입찰업무 정리
    "조달진행상황",          # 조달진행상황 엑셀 파일 정리
    "조달관련그룹웨어공지",   # 조달 관련 그룹웨어 공지
]

# 이 단어로 '시작'하는 업무들은 각각 하나로 합침 (입찰끼리, 조달끼리).
#  제목이 짧거나(입찰) 조금씩 달라도(조달 관련 / 조달 관련 업무) 한 묶음으로.
합치기_접두 = ["입찰", "조달"]

# AI 요약(Claude API) 설정
#  - 키마다 지원 모델이 달라, 앞에서부터 차례로 시도해 되는 모델을 사용
AI_MODELS = [
    "claude-sonnet-4-5",
    "claude-3-7-sonnet-latest",
    "claude-3-5-sonnet-latest",
    "claude-3-5-haiku-latest",
]
AI_MODEL = AI_MODELS[0]
AI_MAX_TOKENS = 6000
AI_KEY_FILE = "claude_api_key.txt"   # exe 옆에 저장(처음 한 번 입력하면 자동 사용)

# 비슷한 업무(다른 구분에 중복) 검출 기준 (0~1).
#  0.30(30%)은 구조만 비슷한 무관 업무까지 과하게 잡혀, 0.55로 설정.
#  더 많이 잡고 싶으면 0.45 등으로 낮추세요.
SIMILARITY_THRESHOLD = 0.55

# ===================================================================
#  스타일 정의
# ===================================================================
FONT_NAME = "맑은 고딕"
색_헤더 = "D9D9D9"     # 표 머리글 / 구분 셀 배경 (연회색)
색_게시판 = "CFD8DC"   # 게시판 현황 머리글 배경

thin = Side(style="thin", color="000000")
ALL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def f(size=10, bold=False, color="000000"):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_MID = Alignment(horizontal="center", vertical="center", wrap_text=True)
FILL_HEADER = PatternFill("solid", fgColor=색_헤더)
FILL_BOARD = PatternFill("solid", fgColor=색_게시판)


# ===================================================================
#  입력(HTML 형식의 .xls) 파싱
# ===================================================================
def _clean_cell(raw):
    s = re.sub(r'(?i)<br\s*/?>', '\n', raw)      # 줄바꿈 보존
    s = re.sub(r'<[^>]+>', '', s)                 # 나머지 태그 제거
    s = html.unescape(s)
    lines = [ln.rstrip() for ln in s.split('\n')]
    return '\n'.join(lines).strip()


def parse_report(path):
    """입력 파일 형식을 자동 판별하여 (업무목록, 게시판목록) 추출.
       - 진짜 엑셀(.xlsx, ZIP=PK 시그니처) → openpyxl 파서
       - HTML 형식(.xls/.htm, NOWSYSTEM 구버전 export) → HTML 파서
    """
    with open(path, 'rb') as fp:
        head = fp.read(4)
    if head[:2] == b'PK':            # .xlsx (ZIP)
        return _parse_xlsx(path)
    return _parse_html(path)


PAD_TOP = 1       # 셀 위 여백(줄)
PAD_BOTTOM = 3    # 셀 아래 여백(줄) — Excel→PDF 시 아래 잘림 방지로 넉넉히


def _pad(s):
    """셀 내용 위·아래로 여백 (빈 칸은 그대로)."""
    s = '' if s is None else str(s)
    return '\n' * PAD_TOP + s + '\n' * PAD_BOTTOM if s.strip() else s


def _is_placeholder(s):
    """ㅇㅇㅇ, ---, (내용 없음) 같은 의미 없는 placeholder 인지 판별."""
    t = re.sub(r'[ㅇoO0○●\s\-·.…()\[\]]', '', s or '')
    return t == '' or t == '내용없음'


def _strip_names(s):
    """게시판 작성자 등으로 들어간 사람 이름 [이동숙] 같은 태그를 제거."""
    if not s:
        return s
    for nm in 이름목록:
        s = re.sub(r'\[\s*' + re.escape(nm) + r'\s*\]\s*', '', s)
    return s


def _strip_dates(s):
    """게시판 내용에서 '날짜_' 접두만 제거 (예: '-6/4_새동초등학교_…' → '-새동초등학교_…').
       그 외 날짜(2026-05-29, 6월 4일, 3/6일 등)는 그대로 둔다."""
    if not s:
        return s
    out = []
    for ln in s.split('\n'):
        ln = re.sub(r'(^|[\s\-])\d{1,2}/\d{1,2}_\s*', r'\1', ln)   # '6/4_' 접두 날짜만 제거
        out.append(ln.rstrip())
    return '\n'.join(out)


def _strip_markers(s):
    """표시용으로 [루틴]·[프로젝트] 표시와 사람 이름 태그를 제거 (수기 보고서 스타일)."""
    if not s:
        return s
    s = re.sub(r'\[루틴\]\s*', '', s)
    s = re.sub(r'\[프로젝트\]\s*', '', s)
    s = _strip_names(s)
    return s.strip()


def _task_rank(item):
    """구분 내 정렬 순서: [루틴]=0(위) → 일반=1 → [신규]=2(아래로 몰기)."""
    진행, 진행률, 예정 = item[0], item[1], item[2]
    진행 = 진행 or ""
    예정 = 예정 or ""
    if "[루틴]" in 진행:
        return 0
    if (진행률 or "").strip() == "신규" or "[신규]" in 진행 or "[신규]" in 예정:
        return 2
    return 1


CLUSTER_THRESHOLD = 0.42   # 같은 구분 안에서 유사 업무를 묶는 기준(0~1). 낮출수록 더 잘 묶임.
# 비교 시 무시할 공통 filler 단어(이게 같다고 묶이면 안 되므로 제거)
_CLUSTER_STOP = ['관련', '업무', '처리', '진행', '작성', '확인', '및', '건',
                 '등록', '신청', '완료', '요청', '대응', '준비', '정리']


def _topic(text):
    """업무 첫 줄에서 머리표·괄호·공백·공통 filler 단어를 떼어 '핵심어'만 남김(비교용)."""
    s = _strip_markers(text or '').split('\n')[0]
    s = re.sub(r'^[·\-\s]+', '', s)
    s = re.sub(r'\([^)]*\)', '', s)          # (100%) 등 괄호 내용 제거
    s = re.sub(r'\d+\s*%', '', s)            # 남은 % 표기
    for w in _CLUSTER_STOP:
        s = s.replace(w, '')
    return re.sub(r'\s+', '', s)


def _cluster(items):
    """유사한 업무끼리 인접하도록 묶음(전이적 그리디). 입력 순서는 최대한 보존."""
    remaining = list(items)
    out = []
    while remaining:
        cluster = [remaining.pop(0)]
        changed = True
        while changed:
            changed = False
            i = 0
            while i < len(remaining):
                ci = _topic(remaining[i][0])
                if len(ci) >= 2 and any(_similarity(_topic(m[0]), ci) >= CLUSTER_THRESHOLD
                                        for m in cluster):
                    cluster.append(remaining.pop(i))
                    changed = True
                else:
                    i += 1
        out.extend(cluster)
    return out


def _arrange(items):
    """구분 내 정렬: [루틴] 위 → 일반 → [신규] 아래, 각 묶음 안에서 유사 업무끼리 군집."""
    groups = {0: [], 1: [], 2: []}
    for it in items:
        groups[_task_rank(it)].append(it)
    return _cluster(groups[0]) + _cluster(groups[1]) + _cluster(groups[2])


def _title(text):
    """업무의 '제목'(첫 줄 핵심)을 비교용으로 정규화.
       머리표·괄호·끝의 '완료/진행중' 제거. 같은 제목끼리 합치기 위함."""
    first = _strip_markers(text or '').split('\n')[0]
    first = re.sub(r'^[·\-\s]+', '', first)
    first = re.sub(r'\([^)]*\)', '', first)              # (100%), (진행중) 등
    first = re.sub(r'(완료|진행중)\s*$', '', first.strip())  # 끝의 완료/진행중
    return re.sub(r'\s+', '', first)


def _title_mergeable(title):
    """제목이 합치기 대상인지 — 대괄호 마커만 있는([상시] 등) 짧은 건 제외."""
    core = re.sub(r'\[[^\]]*\]', '', title)   # [상시] 같은 마커 제거 후 길이 확인
    return len(core) >= 4 or len(title) >= 8


def _merge_key(title):
    """업무를 어떤 묶음으로 합칠지 키 결정.
       - 합치기_접두로 시작하면 그 접두 묶음(입찰끼리/조달끼리)
       - 아니면 같은 제목 묶음 (합치기 대상일 때만)"""
    if not title:
        return None
    for p in 합치기_접두:
        pn = re.sub(r'\s+', '', p)
        if pn and title.startswith(pn):
            return 'P:' + pn
    return 'T:' + title if _title_mergeable(title) else None


def _merge_group(members):
    """같은 제목 업무들을 1개로 합침. members=[[진행,진행률,예정,이미지],...].
       제목은 가장 짧은(깔끔한) 첫 줄, 세부 줄은 모두 모아 중복 제거. 이미지도 합침."""
    firsts = [(m[0] or '').split('\n')[0] for m in members]
    title_line = min(firsts, key=len) if firsts else ''
    detail, seen = [], set()
    for m in members:
        for ln in (m[0] or '').split('\n')[1:]:
            k = re.sub(r'\s+', '', ln)
            if k and k not in seen:
                seen.add(k); detail.append(ln)
    진행 = title_line + ('\n' + '\n'.join(detail) if detail else '')
    progs = [m[1] for m in members if (m[1] or '').strip()]
    진행률 = progs[0] if progs else ''
    plans, ps = [], set()
    for m in members:
        for ln in (m[2] or '').split('\n'):
            k = re.sub(r'\s+', '', ln)
            if k and k not in ps:
                ps.add(k); plans.append(ln)
    imgs = []
    for m in members:
        if len(m) > 3 and m[3]:
            imgs.extend(m[3])
    return [진행, 진행률, '\n'.join(plans), imgs]


def _dedup_plan(진행, 예정, 진행률=''):
    """'금주 예정'이 '지난주 진행'과 동일(앱 자동 복제)이면 비운다.
       단, 100% 완료된 업무만 — 진행중(<100%) 업무의 예정은 '다음 계획'이라 유지!
       (진행중인데 예정을 지우면 누락처럼 보여 재입력 → 업무 누락 발생)"""
    b = re.sub(r'\s+', '', 예정 or '')
    if not b:
        return ""
    done = re.fullmatch(r'100\s*%?', (진행률 or '').strip()) is not None
    if not done:
        return 예정                       # 진행중이면 예정 그대로 유지
    a = re.sub(r'\s+', '', 진행 or '')
    if b == a:
        return ""
    if len(b) >= 15 and (a.startswith(b) or b.startswith(a)):
        return ""
    return 예정


def _similarity(a, b):
    """두 업무 텍스트의 유사도 0.0~1.0 (공백 제거 후 문자열 비교)."""
    na = re.sub(r'\s+', '', a or '')
    nb = re.sub(r'\s+', '', b or '')
    if not na or not nb:
        return 0.0
    return difflib.SequenceMatcher(None, na, nb).ratio()


def _task_disp(t):
    return 구분_매핑.get(t['분류'], t['분류'] or "기타업무")


def _task_text(t):
    return (t.get('진행', '') or '') + ' ' + (t.get('예정', '') or '')


def find_similar_pairs(all_tasks, threshold=SIMILARITY_THRESHOLD):
    """서로 다른 구분에 있는 '동일 업무를 1개로 묶은' 항목들 중
       유사도>=threshold 인 쌍을 찾는다.
       반환: [ (itemA, itemB, sim), ... ]  (item = {'disp','text','ids'})"""
    uniq = {}
    order = []
    for tasks in all_tasks:
        for t in tasks:
            disp = _task_disp(t)
            txt = _task_text(t)
            key = (disp, re.sub(r'\s+', '', txt))
            if key not in uniq:
                uniq[key] = {'disp': disp, 'text': txt, 'ids': set()}
                order.append(key)
            uniq[key]['ids'].add(id(t))
    items = [uniq[k] for k in order]
    pairs = []
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            if items[i]['disp'] == items[j]['disp']:   # 다른 구분만 확인
                continue
            sim = _similarity(items[i]['text'], items[j]['text'])
            if sim >= threshold:
                pairs.append((items[i], items[j], sim))
    pairs.sort(key=lambda x: -x[2])               # 유사도 높은 순
    return pairs


def find_daily_duplicates(all_tasks, threshold=0.62):
    """본문 업무 중 '데일리 업무(상단 고정)'에 이미 있는 항목과 겹치는 것을 찾는다.
       동일 내용 업무는 1개로 묶음. 반환: [{'task','item','ids'}, ...]"""
    seen, order = {}, []
    for tasks in all_tasks:
        for t in tasks:
            first = _strip_markers(t.get('진행', '')).split('\n')[0]
            fn = re.sub(r'\s+', '', first)
            if len(fn) < 3:
                continue
            for item in 데일리_항목:
                inorm = re.sub(r'\s+', '', item)
                if not inorm:
                    continue
                if inorm in fn or fn in inorm or _similarity(first, item) >= threshold:
                    key = re.sub(r'\s+', '', t.get('진행', '') or '')
                    if key not in seen:
                        seen[key] = {'task': t, 'item': item, 'ids': set()}
                        order.append(key)
                    seen[key]['ids'].add(id(t))
                    break
    return [seen[k] for k in order]


def _fmt_progress(v):
    """진행률 셀 값을 '..%' 문자열로 정규화 (숫자/문자 모두 처리)."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        pct = v * 100 if v <= 1 else v
        return f"{round(pct)}%"
    return str(v).strip()


def _parse_xlsx(path):
    """진짜 엑셀(.xlsx) 파일에서 (업무목록, 게시판목록) 추출.
       구조: 1행 제목 / 3행 헤더 / 이후 A열(업무분류) 병합블록 단위로 업무,
             내용은 B열 줄별, 진행률 C열, 예정 D열. 하단 게시판 섹션 별도."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    def cval(r, c):
        v = ws.cell(r, c).value
        if v is None:
            return ""
        return str(v).strip() if not isinstance(v, (int, float)) else v

    maxr = ws.max_row
    # 헤더('업무분류') 행 찾기
    header_row = None
    for r in range(1, maxr + 1):
        if str(ws.cell(r, 1).value).strip() == '업무분류':
            header_row = r
            break
    start = (header_row + 1) if header_row else 2

    # A열에 값이 있는 행 = 블록 시작점
    starts = [r for r in range(start, maxr + 1) if cval(r, 1) != ""]
    starts.append(maxr + 1)  # 보초값

    tasks, board, in_board = [], [], False
    task_ranges = []   # (task_dict, r0, r1) — 이미지 매핑용
    for i in range(len(starts) - 1):
        r0, r1 = starts[i], starts[i + 1] - 1
        a = str(ws.cell(r0, 1).value).strip()

        if '게시판 현황' in a:        # 게시판 섹션 시작(머리글)
            in_board = True
            continue

        # B/D 열을 줄 단위로 모음
        b_lines = [str(ws.cell(r, 2).value).rstrip() for r in range(r0, r1 + 1)
                   if ws.cell(r, 2).value not in (None, "")]
        d_lines = [str(ws.cell(r, 4).value).rstrip() for r in range(r0, r1 + 1)
                   if ws.cell(r, 4).value not in (None, "")]

        if in_board:
            content = "\n".join(b_lines)
            status = ""
            for r in range(r0, r1 + 1):
                if ws.cell(r, 4).value not in (None, ""):
                    status = str(ws.cell(r, 4).value).strip()
                    break
            if (a + content + status).strip():
                board.append((a, content, status))
            continue

        # 진행률: 블록 내 첫 비어있지 않은 C
        prog = ""
        for r in range(r0, r1 + 1):
            if ws.cell(r, 3).value not in (None, ""):
                prog = _fmt_progress(ws.cell(r, 3).value)
                break

        td = {
            '분류': a,
            '진행': "\n".join(b_lines),
            '진행률': prog,
            '예정': "\n".join(d_lines),
            '이미지': [],
        }
        tasks.append(td)
        task_ranges.append((td, r0, r1))

    # 엑셀에 박힌 이미지를 앵커 위치(행)로 해당 업무에 매핑
    for im in getattr(ws, '_images', []):
        try:
            arow = im.anchor._from.row + 1
            try:
                data = im.ref.getvalue()
            except Exception:
                im.ref.seek(0); data = im.ref.read()
            w, h = int(im.width), int(im.height)
        except Exception:
            continue
        if not data:
            continue
        for td, r0, r1 in task_ranges:
            if r0 <= arow <= r1:
                td['이미지'].append({'data': data, 'w': w, 'h': h})
                break

    return tasks, board


def _parse_html(path):
    """HTML 형식(.xls) 입력에서 (업무목록, 게시판목록) 추출."""
    with open(path, encoding='utf-8', errors='replace') as fp:
        htmltxt = fp.read()

    # Excel '웹페이지' 형식(데이터가 별도 폴더에 있는 경우) 처리
    if 'frameset' in htmltxt.lower() and '업무분류' not in htmltxt:
        sub = os.path.splitext(path)[0] + '.files'
        sheet = os.path.join(sub, 'sheet001.htm')
        if os.path.exists(sheet):
            with open(sheet, encoding='utf-8', errors='replace') as fp:
                htmltxt = fp.read()
        else:
            raise ValueError(
                "이 파일은 표준 형식이 아닙니다(Excel 웹페이지 형식).\n"
                "NOWSYSTEM 앱에서 '주간보고 내보내기'로 다시 저장한 .xls 파일을 사용하세요.\n\n"
                + os.path.basename(path))

    rows = re.findall(r'(?is)<tr[^>]*>(.*?)</tr>', htmltxt)
    tasks, board, in_board = [], [], False

    for r in rows:
        cells = re.findall(r'(?is)<t[dh]([^>]*)>(.*?)</t[dh]>', r)
        if not cells:
            continue
        texts = [_clean_cell(c[1]) for c in cells]
        joined = ' '.join(texts)

        if len(cells) == 1 and '게시판 현황' in joined:
            in_board = True
            continue
        if any(t.strip() == '업무분류' for t in texts):   # 머리글 건너뜀
            continue

        if in_board:
            # 게시판 행: [구분, 내용, 상태]  (셀 수가 다를 수 있어 안전 처리)
            label = texts[0] if len(texts) > 0 else ""
            content = texts[1] if len(texts) > 1 else ""
            status = texts[-1] if len(texts) > 2 else ""
            if (label + content + status).strip():
                board.append((label, content, status))
            continue

        if len(cells) >= 4:
            tasks.append({
                '분류': texts[0].strip(),
                '진행': texts[1],
                '진행률': texts[2].strip(),
                '예정': texts[3],
            })
        elif joined.strip():
            # 예외적 행도 누락 없이 보존 (기타로 분류)
            tasks.append({'분류': '기타', '진행': joined, '진행률': '', '예정': ''})

    return tasks, board


def _header_text(path):
    """입력 파일(xlsx/html)에서 머리글 텍스트('금주 진행 업무 (...)' 등)를 추출."""
    try:
        with open(path, 'rb') as fp:
            head = fp.read(4)
        if head[:2] == b'PK':  # xlsx
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            parts = []
            for r in range(1, min(ws.max_row, 6) + 1):
                for c in range(1, 5):
                    v = ws.cell(r, c).value
                    if isinstance(v, str):
                        parts.append(v)
            return " ".join(parts)
        with open(path, encoding='utf-8', errors='replace') as fp:
            return fp.read()
    except Exception:
        return ""


def derive_report_date(paths):
    """'금주 진행 업무 (시작 ~ 종료)'의 종료일 다음 월요일을 보고일로 추정."""
    for path in paths:
        txt = _header_text(path)
        m = re.search(r'금주 진행 업무 \(\d{4}-\d{2}-\d{2}\s*~\s*(\d{4})-(\d{2})-(\d{2})', txt)
        if m:
            try:
                end = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                days = (-end.weekday()) % 7 or 7   # 종료일 이후 첫 월요일
                return end + datetime.timedelta(days=days)
            except Exception:
                pass
    # 추정 실패 시: 이번 주 월요일
    today = datetime.date.today()
    return today - datetime.timedelta(days=today.weekday())


# ===================================================================
#  엑셀(.xlsx) 생성
# ===================================================================
요일 = ['월', '화', '수', '목', '금', '토', '일']


def merge_tasks(all_tasks):
    """입력 업무들을 구분별 버킷으로 병합. 반환 (buckets, extra_order).
       buckets[disp] = [[진행, 진행률, 예정], ...]  (예정 채우기 위해 리스트로)."""
    all_items, extra_order = [], []
    for tasks in all_tasks:
        for t in tasks:
            disp = 구분_매핑.get(t['분류'], t['분류'] or "기타업무")
            if disp not in 구분_매핑.values() and disp not in extra_order:
                extra_order.append(disp)
            예정 = _dedup_plan(t['진행'], t['예정'], t['진행률'])
            all_items.append({'disp': disp, '진행': t['진행'], '진행률': t['진행률'],
                              '예정': 예정, '이미지': t.get('이미지', [])})

    merge_groups = {}
    for i, it in enumerate(all_items):
        k = _merge_key(_title(it['진행']))
        if k:
            merge_groups.setdefault(k, []).append(i)

    buckets, seen_keys, used = {}, {}, set()
    for i, it in enumerate(all_items):
        if i in used:
            continue
        k = _merge_key(_title(it['진행']))
        grp = merge_groups.get(k, [i]) if k else [i]
        if len(grp) > 1:                       # 같은 제목/접두 여러 개 → 합치기
            used.update(grp)
            members = [all_items[j] for j in grp]
            cats = [m['disp'] for m in members]
            disp = max(set(cats), key=cats.count)
            row = _merge_group([[m['진행'], m['진행률'], m['예정'], m['이미지']] for m in members])
        else:
            used.add(i)
            disp = it['disp']
            row = [it['진행'], it['진행률'], it['예정'], it['이미지']]
        key = re.sub(r'\s+', '', row[0] or '') + '||' + re.sub(r'\s+', '', row[2] or '')
        s = seen_keys.setdefault(disp, set())
        if key in s and key != '||':
            continue
        s.add(key)
        buckets.setdefault(disp, []).append(row)
    return buckets, extra_order


def find_missing_plans(buckets):
    """진행률 100% 미만인데 '금주 예정 업무'가 비어있는 업무를 찾는다.
       반환: [{'disp','row'(리스트 참조)}, ...]"""
    needs = []
    for disp, rows in buckets.items():
        for row in rows:
            진행률 = (row[1] or '').strip()
            예정 = (row[2] or '').strip()
            m = re.match(r'(\d+)\s*%$', 진행률)
            if m and int(m.group(1)) < 100 and not 예정:
                needs.append({'disp': disp, 'row': row})
    return needs


def build_workbook(buckets, extra_order, merged_board, report_date):
    """buckets/extra_order = merge_tasks() 결과, merged_board = 통합 게시판."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주간보고"

    # 열 너비 (A:구분/라벨, B:내용, C:진행률/주기, D:금주예정/담당자)
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 30

    R = 1

    def style_row(row, fill=None, border=True):
        for col in range(1, 5):
            c = ws.cell(row=row, column=col)
            if border:
                c.border = ALL_BORDER
            if fill:
                c.fill = fill

    # --- 제목 ---
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
    c = ws.cell(R, 1, f"◎ {부서명} 주간 보고 ◎")
    c.font = f(18, bold=True)
    c.alignment = CENTER
    ws.row_dimensions[R].height = 34
    R += 1

    # --- 보고일 / 보고자 ---
    dstr = f"{report_date.year}.{report_date.month:02d}.{report_date.day:02d}. ({요일[report_date.weekday()]})"
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
    c = ws.cell(R, 1, f"업무 보고일 : {dstr}")
    c.font = f(11); c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[R].height = 20
    R += 1
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
    c = ws.cell(R, 1, f"업무 보고자 :  {보고자}")
    c.font = f(11); c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[R].height = 20
    R += 1

    R += 1  # 빈 줄

    # --- 업무 체크리스트 표 ---
    head_row = R
    ws.cell(R, 2, "업무 내용").font = f(11, bold=True)
    ws.cell(R, 3, "주기").font = f(11, bold=True)
    ws.cell(R, 4, "담당자").font = f(11, bold=True)
    for col in (2, 3, 4):
        ws.cell(R, col).alignment = CENTER
    style_row(R, fill=FILL_HEADER)
    R += 1
    for 내용, 주기, 담당 in 체크리스트:
        ws.cell(R, 2, _pad(내용)).font = f(10); ws.cell(R, 2).alignment = LEFT
        ws.cell(R, 3, 주기).font = f(10); ws.cell(R, 3).alignment = CENTER
        ws.cell(R, 4, 담당).font = f(10); ws.cell(R, 4).alignment = CENTER
        style_row(R)
        R += 1
    # 좌측 '업무 체크리스트' 세로 병합 라벨
    ws.merge_cells(start_row=head_row, start_column=1, end_row=R - 1, end_column=1)
    lc = ws.cell(head_row, 1, "업무\n체크리스트")
    lc.font = f(11, bold=True); lc.alignment = CENTER; lc.fill = FILL_HEADER
    for rr in range(head_row, R):
        ws.cell(rr, 1).border = ALL_BORDER
        ws.cell(rr, 1).fill = FILL_HEADER

    R += 1  # 빈 줄

    # --- 본문 표 머리글 ---
    ws.cell(R, 1, "구분").font = f(11, bold=True)
    ws.cell(R, 2, "지난주 진행 업무").font = f(11, bold=True)
    ws.cell(R, 3, "진행률").font = f(11, bold=True)
    ws.cell(R, 4, "금주 예정 업무").font = f(11, bold=True)
    for col in range(1, 5):
        ws.cell(R, col).alignment = CENTER
    style_row(R, fill=FILL_HEADER)
    R += 1

    def add_image_row(img):
        """업무 아래에 이미지 한 줄 추가 (B열, 폭에 맞춰 축소)."""
        nonlocal R
        if XLImage is None or not img.get('data'):
            return
        try:
            xi = XLImage(io.BytesIO(img['data']))
            maxw = 430
            w0 = img.get('w') or xi.width or maxw
            h0 = img.get('h') or xi.height or 300
            ratio = min(maxw / w0, 1.0)
            xi.width = int(w0 * ratio)
            xi.height = int(h0 * ratio)
            ws.add_image(xi, f"B{R}")
            ws.row_dimensions[R].height = max(xi.height * 0.75 + 4, 20)
            style_row(R)
            R += 1
        except Exception:
            pass

    def write_block(label, items):
        """label = 구분명, items = [[지난주, 진행률, 금주, 이미지?], ...]  세로 병합 처리"""
        nonlocal R
        if not items:
            return
        start = R
        for item in items:
            지난주, 진행률, 금주 = item[0], item[1], item[2]
            images = item[3] if len(item) > 3 else None
            ws.cell(R, 2, _pad(_strip_markers(지난주))).font = f(10); ws.cell(R, 2).alignment = LEFT
            ws.cell(R, 3, 진행률).font = f(10, bold=True); ws.cell(R, 3).alignment = CENTER
            ws.cell(R, 4, _pad(_strip_markers(금주))).font = f(10); ws.cell(R, 4).alignment = LEFT
            style_row(R)
            R += 1
            if images:
                for img in images:
                    add_image_row(img)
        ws.merge_cells(start_row=start, start_column=1, end_row=R - 1, end_column=1)
        lc = ws.cell(start, 1, label)
        lc.font = f(11, bold=True); lc.alignment = CENTER; lc.fill = FILL_HEADER
        for rr in range(start, R):
            ws.cell(rr, 1).border = ALL_BORDER
            ws.cell(rr, 1).fill = FILL_HEADER

    # 데일리 업무 (고정)
    write_block("데일리\n업무", [(데일리_지난주, 데일리_진행률, 데일리_금주)])

    # buckets/extra_order 는 merge_tasks() 에서 미리 병합되어 전달됨

    # 통합 게시판에서 '세금계산서·분납서'를 재무업무 맨 위 고정 행으로 추출
    pinned_top = []
    pin_contents = []   # (키워드, 내용) — 본문 중복 제거 비교용
    for label, content, status in merged_board:
        if any(k in label for k in 재무_고정_게시판):
            txt = f"{label} : {content}" if content.strip() else label
            pinned_top.append((txt, status, ""))
            for k in 재무_고정_게시판:
                if k in label:
                    pin_contents.append((k, content))

    # 게시판 맨위 고정(세금계산서/분납서)과 겹치는 본문 업무는 제거 (중복 방지)
    for disp in list(buckets.keys()):
        kept = []
        for row in buckets[disp]:
            text = row[0] or ''
            dup = any(k in text and _similarity(text, pc) >= 0.45
                      for k, pc in pin_contents)
            if not dup:
                kept.append(row)
        buckets[disp] = kept

    def _pin_index(text):
        n = re.sub(r'\s+', '', text or '')
        for i, k in enumerate(조달입찰_고정_키워드):
            if re.sub(r'\s+', '', k) in n:
                return i
        return -1

    order = list(구분_순서) + [x for x in extra_order if x not in 구분_순서]
    for disp in order:
        rows = []
        if disp == 재무_고정_구분 and pinned_top:
            rows += pinned_top                       # 세금계산서·분납서 맨 위 고정(중복 합침)
        if disp in buckets:
            items = buckets[disp]
            if disp == 조달입찰_고정_구분:
                # 표준 업무(키워드 일치)를 키워드 순서대로 맨 위로 고정, 나머지는 군집 정렬
                pinned = sorted([it for it in items if _pin_index(it[0]) >= 0],
                                key=lambda it: _pin_index(it[0]))
                rest = [it for it in items if _pin_index(it[0]) < 0]
                rows += pinned + _arrange(rest)
            else:
                # 구분 내 정렬: [루틴] 위 → 일반 → [신규] 아래, 유사 업무끼리 군집
                rows += _arrange(items)
        if rows:
            write_block(disp, rows)

    # --- 부서 공용 게시판 현황 (재무로 올린 항목 제외) ---
    bottom = [(l, c, s) for (l, c, s) in merged_board
              if not any(k in l for k in 재무_고정_게시판)]
    if bottom:
        ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
        c = ws.cell(R, 1, "📂 부서 공용 게시판 현황")
        c.font = f(11, bold=True); c.alignment = CENTER; c.fill = FILL_BOARD
        style_row(R, fill=FILL_BOARD)
        R += 1
        for label, content, status in bottom:
            ws.cell(R, 1, label).font = f(10, bold=True)
            ws.cell(R, 1).alignment = CENTER
            ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=3)
            ws.cell(R, 2, _pad(content)).font = f(10); ws.cell(R, 2).alignment = LEFT
            ws.cell(R, 4, status).font = f(10, bold=True); ws.cell(R, 4).alignment = CENTER
            style_row(R)
            R += 1

    # --- 인쇄 설정 (A4 세로, 가로 1페이지 맞춤) ---
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.print_options.horizontalCentered = True

    return wb


# ===================================================================
#  AI 요약 (Claude API)
# ===================================================================
def app_dir():
    """exe(또는 스크립트)가 있는 폴더 — API 키 파일 위치."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def load_api_key():
    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if key:
        return key
    path = os.path.join(app_dir(), AI_KEY_FILE)
    if os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as fp:
                return fp.read().strip()
        except Exception:
            pass
    return ""


def save_api_key(key):
    try:
        with open(os.path.join(app_dir(), AI_KEY_FILE), "w", encoding="utf-8") as fp:
            fp.write(key.strip())
    except Exception:
        pass


def _tasks_as_text(all_tasks):
    """모든 업무를 구분별로 묶어 요약용 텍스트로 직렬화."""
    lines = []
    for tasks in all_tasks:
        for t in tasks:
            분류 = t.get("분류", "")
            진행 = (t.get("진행", "") or "").replace("\n", " ")
            진행률 = t.get("진행률", "")
            예정 = (t.get("예정", "") or "").replace("\n", " ")
            seg = f"[{분류}] ({진행률}) {진행}"
            if 예정.strip() and 예정.strip() != 진행.strip():
                seg += f"  | 예정: {예정}"
            lines.append(seg)
    return "\n".join(lines)


def generate_ai_summary(all_tasks, report_date, api_key):
    """Claude API로 두 가지 요약(구분별 핵심 / 완료·진행중·예정) 생성 → dict 반환."""
    body_text = _tasks_as_text(all_tasks)
    prompt = (
        f"다음은 {부서명}의 {report_date.strftime('%Y-%m-%d')} 주간 업무 목록입니다. "
        "각 줄은 [업무분류] (진행률) 업무내용 형식입니다.\n\n"
        f"{body_text}\n\n"
        "위 내용을 바탕으로 임원 보고용 주간 요약을 작성하세요. "
        "주요 업무가 누락되지 않게 하되 문장은 간결하게 다듬으세요. "
        "반드시 아래 JSON 형식만 출력하세요(코드블록·설명 없이 순수 JSON):\n"
        "{\n"
        '  "by_category": [ {"구분": "경영관리", "items": ["핵심 업무 한 줄 (진행률)", "..."]}, ... ],\n'
        '  "by_status": { "완료": ["..."], "진행중": ["...(진행률)"], "예정": ["..."] }\n'
        "}\n"
        "- by_category 구분명은 입력의 업무분류를 자연스럽게 묶어서 사용.\n"
        "- 완료=100%인 업무, 진행중=1~99%, 예정=신규/다음주 계획.\n"
        "- 각 항목은 한국어 한 줄."
    )
    # 지원 모델을 순서대로 시도 (키마다 가능한 모델이 다름)
    last_err = None
    for model in AI_MODELS:
        payload = json.dumps({
            "model": model,
            "max_tokens": AI_MAX_TOKENS,
            "messages": [{"role": "user", "content": prompt}],
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "content-type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            try:
                emsg = json.loads(e.read().decode("utf-8", "replace")).get("error", {}).get("message", "")
            except Exception:
                emsg = ""
            e._apimsg = emsg
            # '모델을 못 찾음'류만 다음 모델로 재시도, 그 외(크레딧/키 등)는 즉시 중단
            if e.code in (400, 404) and ("model" in emsg.lower() and "not" in emsg.lower()):
                last_err = e
                continue
            raise
        text = "".join(blk.get("text", "") for blk in data.get("content", []))
        m = re.search(r"\{.*\}", text, re.S)
        if not m:
            raise ValueError("AI 응답을 해석할 수 없습니다.")
        return json.loads(m.group(0))
    if last_err is not None:
        raise last_err
    raise ValueError("사용 가능한 AI 모델을 찾지 못했습니다.")


def build_summary_workbook(summary, report_date):
    """AI 요약 dict → 보기 좋은 .xlsx 워크북."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI요약"
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 86
    R = 1

    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=2)
    c = ws.cell(R, 1, f"◎ {부서명} 주간 보고 — AI 요약 ◎")
    c.font = f(16, bold=True); c.alignment = CENTER
    ws.row_dimensions[R].height = 30
    R += 1
    dstr = f"{report_date.year}.{report_date.month:02d}.{report_date.day:02d}. ({요일[report_date.weekday()]})"
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=2)
    c = ws.cell(R, 1, f"보고일 : {dstr}   |   보고자 : {보고자}")
    c.font = f(10); c.alignment = Alignment(horizontal="left", vertical="center")
    R += 2

    def section(title, fill):
        nonlocal R
        ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=2)
        c = ws.cell(R, 1, title)
        c.font = f(12, bold=True, color="FFFFFF"); c.alignment = CENTER
        c.fill = PatternFill("solid", fgColor=fill)
        for col in (1, 2):
            ws.cell(R, col).border = ALL_BORDER
        R += 1

    def row(label, text):
        nonlocal R
        a = ws.cell(R, 1, label); a.font = f(10, bold=True); a.alignment = CENTER; a.fill = FILL_HEADER
        b = ws.cell(R, 2, text); b.font = f(10); b.alignment = LEFT
        for col in (1, 2):
            ws.cell(R, col).border = ALL_BORDER
        R += 1

    # 1) 구분별 핵심 요약
    section("Ⅰ. 구분별 핵심 요약", "1f3b57")
    for grp in summary.get("by_category", []):
        구분 = grp.get("구분", "")
        items = grp.get("items", [])
        row(구분, "\n".join(f"· {it}" for it in items))

    R += 1
    # 2) 완료 / 진행중 / 예정
    section("Ⅱ. 완료 · 진행중 · 예정", "1f9d55")
    bs = summary.get("by_status", {})
    for key, color in (("완료", "2e7d32"), ("진행중", "e08a00"), ("예정", "1565c0")):
        items = bs.get(key, [])
        a = ws.cell(R, 1, key); a.font = f(11, bold=True, color="FFFFFF"); a.alignment = CENTER
        a.fill = PatternFill("solid", fgColor=color)
        b = ws.cell(R, 2, "\n".join(f"· {it}" for it in items) if items else "-")
        b.font = f(10); b.alignment = LEFT
        for col in (1, 2):
            ws.cell(R, col).border = ALL_BORDER
        R += 1

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    return wb


# ===================================================================
#  GUI 프로그램 (눈에 보이는 메인 창)
# ===================================================================
from tkinter import ttk


def _short(text, n=110):
    s = re.sub(r'\s*\n\s*', ' / ', (text or '').strip())
    return s if len(s) <= n else s[:n] + " …"


def resolve_similar_pairs(parent, pairs):
    """비슷한 업무 쌍을 팝업으로 보여주고 합치기/개별 결정을 받는다.
       반환: 제거할 업무들의 id 집합 (취소 시 None)."""
    dlg = tk.Toplevel(parent)
    dlg.title("비슷한 업무 확인")
    dlg.configure(bg="#f4f6f8")
    dlg.transient(parent)
    dlg.grab_set()
    w, h = 820, 600
    x = parent.winfo_rootx() + 40
    y = parent.winfo_rooty() + 20
    dlg.geometry(f"{w}x{h}+{x}+{y}")
    dlg.attributes("-topmost", True)

    tk.Label(dlg, text="다른 구분에 비슷한 업무가 있습니다",
             font=(FONT_NAME, 14, "bold"), bg="#f4f6f8", fg="#1f3b57").pack(pady=(12, 2))
    tk.Label(dlg, text=f"총 {len(pairs)}쌍 — 같은 업무면 '합치기', 다른 업무면 '개별로 두기'를 고르세요.",
             font=(FONT_NAME, 10), bg="#f4f6f8", fg="#555").pack(pady=(0, 8))

    # 스크롤 영역
    mid = tk.Frame(dlg, bg="#f4f6f8")
    mid.pack(fill="both", expand=True, padx=12)
    canvas = tk.Canvas(mid, bg="#f4f6f8", highlightthickness=0)
    sb = ttk.Scrollbar(mid, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas, bg="#f4f6f8")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw", width=w - 40)
    canvas.configure(yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True)
    sb.pack(side="right", fill="y")
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

    choices = []   # 각 쌍의 tk.IntVar (0=개별, 1=A만, 2=B만)
    for idx, (ia, ib, sim) in enumerate(pairs):
        da, db = ia['disp'], ib['disp']
        lf = tk.LabelFrame(inner, text=f"  유사도 {round(sim*100)}%  ",
                           font=(FONT_NAME, 10, "bold"), bg="#ffffff", fg="#c0392b",
                           padx=10, pady=6)
        lf.pack(fill="x", pady=6, padx=2)
        tk.Label(lf, text=f"Ⓐ [{da}]  {_short(ia['text'])}", font=(FONT_NAME, 9),
                 bg="#ffffff", fg="#222", justify="left", anchor="w", wraplength=w - 90).pack(fill="x")
        tk.Label(lf, text=f"Ⓑ [{db}]  {_short(ib['text'])}", font=(FONT_NAME, 9),
                 bg="#ffffff", fg="#222", justify="left", anchor="w", wraplength=w - 90).pack(fill="x", pady=(0, 4))
        var = tk.IntVar(value=0)
        choices.append(var)
        rrow = tk.Frame(lf, bg="#ffffff"); rrow.pack(fill="x")
        tk.Radiobutton(rrow, text="개별로 둘 다 유지", variable=var, value=0,
                       font=(FONT_NAME, 9), bg="#ffffff").pack(side="left")
        tk.Radiobutton(rrow, text=f"합치기 — Ⓐ[{da}]만 남김", variable=var, value=1,
                       font=(FONT_NAME, 9), bg="#ffffff").pack(side="left", padx=10)
        tk.Radiobutton(rrow, text=f"합치기 — Ⓑ[{db}]만 남김", variable=var, value=2,
                       font=(FONT_NAME, 9), bg="#ffffff").pack(side="left")

    result = {"drop": None}

    def on_ok():
        drop = set()
        for (ia, ib, sim), var in zip(pairs, choices):
            if var.get() == 1:
                drop |= ib['ids']
            elif var.get() == 2:
                drop |= ia['ids']
        result["drop"] = drop
        canvas.unbind_all("<MouseWheel>")
        dlg.destroy()

    def on_cancel():
        result["drop"] = None
        canvas.unbind_all("<MouseWheel>")
        dlg.destroy()

    btns = tk.Frame(dlg, bg="#f4f6f8"); btns.pack(fill="x", padx=12, pady=10)
    tk.Button(btns, text="✓ 이 결과로 보고서 만들기", font=(FONT_NAME, 11, "bold"),
              bg="#1f9d55", fg="white", relief="flat", padx=14, pady=6,
              command=on_ok).pack(side="right")
    tk.Button(btns, text="취소", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6, command=on_cancel).pack(side="right", padx=8)
    tk.Button(btns, text="전부 개별 유지", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6,
              command=lambda: [v.set(0) for v in choices]).pack(side="left")
    tk.Button(btns, text="전부 합치기(Ⓐ남김)", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6,
              command=lambda: [v.set(1) for v in choices]).pack(side="left", padx=8)

    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    parent.wait_window(dlg)
    return result["drop"]


def resolve_daily_duplicates(parent, candidates):
    """'데일리 업무'와 겹치는 본문 업무들을 보여주고 뺄지 확인.
       반환: 제거할 업무 id 집합 (취소 시 None)."""
    dlg = tk.Toplevel(parent)
    dlg.title("데일리 업무 중복 확인")
    dlg.configure(bg="#f4f6f8")
    dlg.transient(parent); dlg.grab_set()
    w, h = 820, 560
    dlg.geometry(f"{w}x{h}+{parent.winfo_rootx()+40}+{parent.winfo_rooty()+20}")
    dlg.attributes("-topmost", True)

    tk.Label(dlg, text="'데일리 업무'와 겹치는 본문 업무",
             font=(FONT_NAME, 14, "bold"), bg="#f4f6f8", fg="#1f3b57").pack(pady=(12, 2))
    tk.Label(dlg, text=f"아래 업무는 상단 '데일리 업무'에 이미 있습니다. 본문에서 뺄지 선택하세요. (총 {len(candidates)}건)",
             font=(FONT_NAME, 10), bg="#f4f6f8", fg="#555").pack(pady=(0, 8))

    mid = tk.Frame(dlg, bg="#f4f6f8"); mid.pack(fill="both", expand=True, padx=12)
    canvas = tk.Canvas(mid, bg="#f4f6f8", highlightthickness=0)
    sb = ttk.Scrollbar(mid, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas, bg="#f4f6f8")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw", width=w - 40)
    canvas.configure(yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

    choices = []
    for c in candidates:
        t, item = c['task'], c['item']
        lf = tk.LabelFrame(inner, text=f"  데일리 항목: {item}  ", font=(FONT_NAME, 10, "bold"),
                           bg="#ffffff", fg="#c0392b", padx=10, pady=6)
        lf.pack(fill="x", pady=5, padx=2)
        tk.Label(lf, text=f"본문 업무: {_short(_strip_markers(t.get('진행','')), 95)}  [{_task_disp(t)}]",
                 font=(FONT_NAME, 9), bg="#ffffff", fg="#222",
                 justify="left", anchor="w", wraplength=w - 90).pack(fill="x", pady=(0, 4))
        var = tk.IntVar(value=1)   # 기본: 빼기
        choices.append((var, c))
        rrow = tk.Frame(lf, bg="#ffffff"); rrow.pack(fill="x")
        tk.Radiobutton(rrow, text="본문에서 빼기 (데일리에 이미 있음)", variable=var, value=1,
                       font=(FONT_NAME, 9), bg="#ffffff").pack(side="left")
        tk.Radiobutton(rrow, text="그대로 두기", variable=var, value=0,
                       font=(FONT_NAME, 9), bg="#ffffff").pack(side="left", padx=12)

    result = {"drop": None}

    def on_ok():
        drop = set()
        for var, c in choices:
            if var.get() == 1:
                drop |= c['ids']
        result["drop"] = drop
        canvas.unbind_all("<MouseWheel>"); dlg.destroy()

    def on_cancel():
        result["drop"] = None
        canvas.unbind_all("<MouseWheel>"); dlg.destroy()

    btns = tk.Frame(dlg, bg="#f4f6f8"); btns.pack(fill="x", padx=12, pady=10)
    tk.Button(btns, text="✓ 이 결과로 진행", font=(FONT_NAME, 11, "bold"),
              bg="#1f9d55", fg="white", relief="flat", padx=14, pady=6,
              command=on_ok).pack(side="right")
    tk.Button(btns, text="취소", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6, command=on_cancel).pack(side="right", padx=8)
    tk.Button(btns, text="전부 빼기", font=(FONT_NAME, 10), relief="groove", padx=12, pady=6,
              command=lambda: [v.set(1) for v, _ in choices]).pack(side="left")
    tk.Button(btns, text="전부 두기", font=(FONT_NAME, 10), relief="groove", padx=12, pady=6,
              command=lambda: [v.set(0) for v, _ in choices]).pack(side="left", padx=8)

    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    parent.wait_window(dlg)
    return result["drop"]


def resolve_missing_plans(parent, needs):
    """진행률 100% 미만인데 금주 예정이 빈 업무에 무엇을 넣을지 입력받음.
       입력한 값은 해당 row[2]에 채움. 반환 True(진행)/None(취소)."""
    dlg = tk.Toplevel(parent)
    dlg.title("금주 예정 업무 입력")
    dlg.configure(bg="#f4f6f8")
    dlg.transient(parent); dlg.grab_set()
    w, h = 860, 600
    dlg.geometry(f"{w}x{h}+{parent.winfo_rootx()+30}+{parent.winfo_rooty()+15}")
    dlg.attributes("-topmost", True)

    tk.Label(dlg, text="금주 예정 업무 입력",
             font=(FONT_NAME, 14, "bold"), bg="#f4f6f8", fg="#1f3b57").pack(pady=(12, 2))
    tk.Label(dlg, text=f"진행률이 100% 미만인데 '금주 예정 업무'가 비어있는 업무가 {len(needs)}건 있습니다.\n"
                       "무엇을 넣을지 입력하세요. (비워두면 그대로 빈칸으로 둡니다)",
             font=(FONT_NAME, 10), bg="#f4f6f8", fg="#555", justify="center").pack(pady=(0, 8))

    mid = tk.Frame(dlg, bg="#f4f6f8"); mid.pack(fill="both", expand=True, padx=12)
    canvas = tk.Canvas(mid, bg="#f4f6f8", highlightthickness=0)
    sb = ttk.Scrollbar(mid, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas, bg="#f4f6f8")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw", width=w - 40)
    canvas.configure(yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

    boxes = []
    for n in needs:
        row = n['row']
        lf = tk.LabelFrame(inner, text=f"  [{n['disp']}]  ",
                           font=(FONT_NAME, 10, "bold"), bg="#ffffff", fg="#c0392b",
                           padx=10, pady=6)
        lf.pack(fill="x", pady=5, padx=2)
        # 지난주 진행 업무 — 전체 내용 표시
        tk.Label(lf, text="지난주 진행 업무 (전체):", font=(FONT_NAME, 9, "bold"),
                 bg="#ffffff", fg="#555", anchor="w").pack(fill="x")
        tk.Label(lf, text=_strip_markers(row[0]), font=(FONT_NAME, 9),
                 bg="#f7f9fb", fg="#222", justify="left", anchor="w",
                 wraplength=w - 80).pack(fill="x", pady=(0, 6))
        # 진행률 — 수정 가능
        prow = tk.Frame(lf, bg="#ffffff"); prow.pack(fill="x")
        tk.Label(prow, text="진행률:", font=(FONT_NAME, 9, "bold"),
                 bg="#ffffff").pack(side="left")
        pent = tk.Entry(prow, width=8, font=(FONT_NAME, 10), justify="center")
        pent.insert(0, row[1] or "")
        pent.pack(side="left", padx=6)
        tk.Label(prow, text="← 다 끝났으면 100% 로 바꾸세요 (그러면 예정 안 써도 됨)",
                 font=(FONT_NAME, 9), bg="#ffffff", fg="#888").pack(side="left")
        # 금주 예정 — 직접 입력
        tk.Label(lf, text="금주 예정 업무 ↓ (직접 입력)", font=(FONT_NAME, 9, "bold"),
                 bg="#ffffff", fg="#1f3b57", anchor="w").pack(fill="x", pady=(6, 0))
        txt = tk.Text(lf, height=2, font=(FONT_NAME, 10), wrap="word", relief="solid", bd=1)
        txt.pack(fill="x")
        boxes.append((pent, txt, row))

    result = {"ok": None}

    def on_ok():
        for pent, txt, row in boxes:
            pv = pent.get().strip()
            if pv:
                if re.fullmatch(r'\d+', pv):
                    pv += '%'
                row[1] = pv
            val = txt.get("1.0", "end").strip()
            if val:
                row[2] = val
        result["ok"] = True
        canvas.unbind_all("<MouseWheel>"); dlg.destroy()

    def on_cancel():
        result["ok"] = None
        canvas.unbind_all("<MouseWheel>"); dlg.destroy()

    btns = tk.Frame(dlg, bg="#f4f6f8"); btns.pack(fill="x", padx=12, pady=10)
    tk.Button(btns, text="✓ 입력한 대로 진행", font=(FONT_NAME, 11, "bold"),
              bg="#1f9d55", fg="white", relief="flat", padx=14, pady=6,
              command=on_ok).pack(side="right")
    tk.Button(btns, text="모두 비워두고 진행", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6, command=on_ok).pack(side="right", padx=8)
    tk.Button(btns, text="취소", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6, command=on_cancel).pack(side="left")

    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    parent.wait_window(dlg)
    return result["ok"]


def resolve_board(parent, all_boards, filenames):
    """게시판 항목을 라벨별로 1개로 통합. 파일마다 내용이 다르면 팝업으로 선택.
       반환: [(label, content, status), ...]  또는 None(취소)."""
    groups, order = {}, []
    for fi, board in enumerate(all_boards):
        for label, content, status in board:
            lbl = _strip_names(label).strip()
            con = _strip_dates(_strip_names(content)).strip()   # 게시판 내용은 날짜 제거
            if _is_placeholder(con):
                con = ''
            key = re.sub(r'\s+', '', lbl)
            if key not in groups:
                groups[key] = {'label': lbl, 'opts': []}
                order.append(key)
            g = groups[key]
            match = None
            for o in g['opts']:
                if re.sub(r'\s+', '', o['content']) == re.sub(r'\s+', '', con):
                    match = o
                    break
            if match:
                if filenames[fi] not in match['files']:
                    match['files'].append(filenames[fi])
                if status and not match['status']:
                    match['status'] = status
            else:
                g['opts'].append({'content': con, 'status': status, 'files': [filenames[fi]]})

    resolved, conflicts = [], []
    for key in order:
        g = groups[key]
        nonempty = [o for o in g['opts'] if o['content']]
        if len(nonempty) <= 1:
            o = nonempty[0] if nonempty else g['opts'][0]
            resolved.append([g['label'], o['content'], o['status'], None])
        else:
            resolved.append([g['label'], None, '', key])
            conflicts.append((key, g['label'], nonempty))

    if conflicts:
        choices = _board_conflict_dialog(parent, conflicts)
        if choices is None:
            return None
        cmap = {}
        for (key, label, opts), ch in zip(conflicts, choices):
            if ch == -1:   # 모두 합치기
                content = "\n".join(o['content'] for o in opts)
                status = next((o['status'] for o in opts if o['status']), '')
            else:
                content, status = opts[ch]['content'], opts[ch]['status']
            cmap[key] = (content, status)
        for row in resolved:
            if row[3] in cmap:
                row[1], row[2] = cmap[row[3]]

    return [(r[0], r[1] or '', r[2]) for r in resolved]


def _board_conflict_dialog(parent, conflicts):
    """파일마다 다른 게시판 항목들을 보여주고 어느 내용을 반영할지 선택."""
    dlg = tk.Toplevel(parent)
    dlg.title("부서 공용 게시판 - 내용 선택")
    dlg.configure(bg="#f4f6f8")
    dlg.transient(parent); dlg.grab_set()
    w, h = 820, 600
    dlg.geometry(f"{w}x{h}+{parent.winfo_rootx()+40}+{parent.winfo_rooty()+20}")
    dlg.attributes("-topmost", True)

    tk.Label(dlg, text="부서 공용 게시판 - 어느 내용을 반영할까요?",
             font=(FONT_NAME, 14, "bold"), bg="#f4f6f8", fg="#1f3b57").pack(pady=(12, 2))
    tk.Label(dlg, text=f"파일마다 내용이 다른 게시판 항목이 {len(conflicts)}개 있습니다. 항목별로 하나를 고르세요.",
             font=(FONT_NAME, 10), bg="#f4f6f8", fg="#555").pack(pady=(0, 8))

    mid = tk.Frame(dlg, bg="#f4f6f8"); mid.pack(fill="both", expand=True, padx=12)
    canvas = tk.Canvas(mid, bg="#f4f6f8", highlightthickness=0)
    sb = ttk.Scrollbar(mid, orient="vertical", command=canvas.yview)
    inner = tk.Frame(canvas, bg="#f4f6f8")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner, anchor="nw", width=w - 40)
    canvas.configure(yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
    canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

    choices = []
    for key, label, opts in conflicts:
        lf = tk.LabelFrame(inner, text=f"  {label}  ", font=(FONT_NAME, 11, "bold"),
                           bg="#ffffff", fg="#1f3b57", padx=10, pady=6)
        lf.pack(fill="x", pady=6, padx=2)
        var = tk.IntVar(value=0)
        choices.append(var)
        for i, o in enumerate(opts):
            who = ", ".join(o['files'])
            tk.Radiobutton(lf, text=f"[{who}]  {_short(o['content'], 90)}",
                           variable=var, value=i, font=(FONT_NAME, 9), bg="#ffffff",
                           justify="left", anchor="w", wraplength=w - 90).pack(fill="x")
        tk.Radiobutton(lf, text="↪ 모두 합치기 (여러 줄로 전부 표시)", variable=var, value=-1,
                       font=(FONT_NAME, 9, "bold"), bg="#ffffff", fg="#1f9d55").pack(fill="x")

    result = {"v": None}

    def on_ok():
        result["v"] = [v.get() for v in choices]
        canvas.unbind_all("<MouseWheel>"); dlg.destroy()

    def on_cancel():
        result["v"] = None
        canvas.unbind_all("<MouseWheel>"); dlg.destroy()

    btns = tk.Frame(dlg, bg="#f4f6f8"); btns.pack(fill="x", padx=12, pady=10)
    tk.Button(btns, text="✓ 이 내용으로 진행", font=(FONT_NAME, 11, "bold"),
              bg="#1f9d55", fg="white", relief="flat", padx=14, pady=6,
              command=on_ok).pack(side="right")
    tk.Button(btns, text="취소", font=(FONT_NAME, 10), relief="groove",
              padx=12, pady=6, command=on_cancel).pack(side="right", padx=8)

    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    parent.wait_window(dlg)
    return result["v"]


class ReportApp:
    def __init__(self, root):
        self.root = root
        self.paths = []

        root.title(f"{부서명} 주간보고서 만들기")
        root.geometry("680x520")
        root.minsize(600, 460)
        root.configure(bg="#f4f6f8")

        # 화면 중앙 배치 + 처음에 맨 앞으로
        root.update_idletasks()
        w, h = 680, 520
        x = (root.winfo_screenwidth() - w) // 2
        y = (root.winfo_screenheight() - h) // 3
        root.geometry(f"{w}x{h}+{x}+{y}")
        root.attributes("-topmost", True)
        root.after(400, lambda: root.attributes("-topmost", False))
        root.lift()
        root.focus_force()

        pad = {"padx": 16}

        # 제목
        tk.Label(root, text=f"◎ {부서명} 주간 보고서 만들기 ◎",
                 font=(FONT_NAME, 16, "bold"), bg="#f4f6f8", fg="#1f3b57"
                 ).pack(pady=(16, 4))
        tk.Label(root, text="3명의 주간보고 엑셀(.xls)을 넣으면 PDF 양식의 통합 엑셀이 만들어집니다.",
                 font=(FONT_NAME, 10), bg="#f4f6f8", fg="#555").pack(pady=(0, 12))

        # 1단계: 파일 선택
        box1 = tk.LabelFrame(root, text=" ① 주간보고 엑셀 선택 (3개) ",
                             font=(FONT_NAME, 11, "bold"), bg="#ffffff", fg="#1f3b57",
                             padx=12, pady=10)
        box1.pack(fill="both", expand=True, **pad)

        btnrow = tk.Frame(box1, bg="#ffffff")
        btnrow.pack(fill="x")
        tk.Button(btnrow, text="📂 파일 추가/선택", font=(FONT_NAME, 11, "bold"),
                  bg="#2d6cdf", fg="white", relief="flat", padx=14, pady=6,
                  command=self.pick_files).pack(side="left")
        tk.Button(btnrow, text="목록 비우기", font=(FONT_NAME, 10),
                  relief="groove", padx=10, pady=6,
                  command=self.clear_files).pack(side="left", padx=8)
        self.count_lbl = tk.Label(btnrow, text="선택된 파일: 0개",
                                  font=(FONT_NAME, 10, "bold"), bg="#ffffff", fg="#2d6cdf")
        self.count_lbl.pack(side="right")

        droptip = ("  ↓  엑셀 파일을 이 칸으로 끌어다 놓으세요 (드래그&드롭)  ↓"
                   if windnd else "  파일 추가/선택 버튼으로 엑셀을 추가하세요")
        self.drop_hint = tk.Label(box1, text=droptip, font=(FONT_NAME, 9, "bold"),
                                  bg="#eef3fb", fg="#2d6cdf", pady=4)
        self.drop_hint.pack(fill="x", pady=(8, 0))

        self.listbox = tk.Listbox(box1, height=6, font=(FONT_NAME, 9),
                                  activestyle="none", bg="#fbfdff")
        self.listbox.pack(fill="both", expand=True, pady=(6, 0))

        # 2단계: 보고일
        box2 = tk.Frame(root, bg="#f4f6f8")
        box2.pack(fill="x", **pad, pady=(12, 4))
        tk.Label(box2, text="② 보고일(월요일):", font=(FONT_NAME, 11, "bold"),
                 bg="#f4f6f8").pack(side="left")
        self.date_var = tk.StringVar(value="")
        tk.Entry(box2, textvariable=self.date_var, font=(FONT_NAME, 12),
                 width=14, justify="center").pack(side="left", padx=8)
        tk.Label(box2, text="(예: 2026-06-08, 파일에서 자동 추정)",
                 font=(FONT_NAME, 9), bg="#f4f6f8", fg="#888").pack(side="left")

        # 옵션
        box3 = tk.Frame(root, bg="#f4f6f8")
        box3.pack(fill="x", **pad, pady=(2, 0))
        self.sim_var = tk.BooleanVar(value=True)
        tk.Checkbutton(box3, text="🔍 중복 확인 (데일리 업무 겹침 · 다른 구분 비슷한 업무) 팝업",
                       variable=self.sim_var, font=(FONT_NAME, 10), bg="#f4f6f8",
                       activebackground="#f4f6f8").pack(side="left")

        box4 = tk.Frame(root, bg="#f4f6f8")
        box4.pack(fill="x", **pad, pady=(0, 4))
        self.ai_var = tk.BooleanVar(value=False)
        tk.Checkbutton(box4, text="🤖 AI 요약본도 함께 생성 (Claude API · 인터넷 필요)",
                       variable=self.ai_var, font=(FONT_NAME, 10), bg="#f4f6f8",
                       activebackground="#f4f6f8").pack(side="left")
        tk.Button(box4, text="API 키 설정", font=(FONT_NAME, 9), relief="groove",
                  command=self.set_api_key).pack(side="right")

        # 3단계: 생성 버튼
        tk.Button(root, text="③  보고서 만들기  ▶", font=(FONT_NAME, 13, "bold"),
                  bg="#1f9d55", fg="white", relief="flat", pady=10,
                  command=self.generate).pack(fill="x", **pad, pady=(8, 4))

        self.status = tk.Label(root, text="준비됨. 파일을 선택하거나 끌어다 놓으세요.",
                               font=(FONT_NAME, 9), bg="#f4f6f8", fg="#555", anchor="w")
        self.status.pack(fill="x", **pad, pady=(0, 12))

        # 드래그&드롭 등록 (창 전체 + 목록칸)
        if windnd:
            root.update()
            try:
                windnd.hook_dropfiles(root, func=self.on_drop, force_unicode=True)
                windnd.hook_dropfiles(self.listbox, func=self.on_drop, force_unicode=True)
            except Exception:
                pass

    # ---- 동작 ----
    def set_status(self, text, color="#555"):
        self.status.config(text=text, fg=color)
        self.root.update_idletasks()

    def on_drop(self, items):
        """끌어다 놓은 파일들을 목록에 추가."""
        exts = (".xls", ".xlsx", ".xlsm", ".htm", ".html")
        added = 0
        for raw in items:
            p = raw.decode("mbcs", "replace") if isinstance(raw, bytes) else str(raw)
            p = p.strip().strip('"')
            if os.path.isdir(p):   # 폴더면 내부 엑셀들 추가
                for name in sorted(os.listdir(p)):
                    fp = os.path.join(p, name)
                    if name.lower().endswith(exts) and fp not in self.paths:
                        self.paths.append(fp); added += 1
                continue
            if p and p not in self.paths:
                self.paths.append(p); added += 1
        if added:
            self.refresh_list()
            if not self.date_var.get().strip():
                self.date_var.set(derive_report_date(self.paths).strftime("%Y-%m-%d"))
            self.set_status(f"{added}개 추가됨 (드래그&드롭). 현재 {len(self.paths)}개.", "#1f3b57")

    def refresh_list(self):
        self.listbox.delete(0, "end")
        for p in self.paths:
            self.listbox.insert("end", "  " + os.path.basename(p))
        self.count_lbl.config(text=f"선택된 파일: {len(self.paths)}개")

    def pick_files(self):
        new = filedialog.askopenfilenames(
            parent=self.root,
            title="주간보고 엑셀을 선택하세요 (Ctrl/Shift로 여러 개)",
            filetypes=[("주간보고 파일", "*.xls *.xlsx *.htm *.html"), ("모든 파일", "*.*")],
        )
        for p in new:
            if p not in self.paths:
                self.paths.append(p)
        self.refresh_list()
        if self.paths and not self.date_var.get().strip():
            self.date_var.set(derive_report_date(self.paths).strftime("%Y-%m-%d"))
        self.set_status(f"{len(self.paths)}개 선택됨.", "#1f3b57")

    def clear_files(self):
        self.paths = []
        self.refresh_list()
        self.set_status("목록을 비웠습니다.")

    def generate(self):
        if not self.paths:
            messagebox.showwarning("파일 없음", "먼저 주간보고 엑셀을 선택하세요.", parent=self.root)
            return
        if len(self.paths) != 3:
            if not messagebox.askyesno("확인",
                    f"{len(self.paths)}개의 파일을 선택했습니다. (보통 3명 = 3개)\n이대로 진행할까요?",
                    parent=self.root):
                return

        # 보고일
        ds = self.date_var.get().strip()
        try:
            report_date = datetime.datetime.strptime(ds, "%Y-%m-%d").date()
        except ValueError:
            report_date = derive_report_date(self.paths)
            self.date_var.set(report_date.strftime("%Y-%m-%d"))

        # 파싱 (읽을 수 없는 파일은 건너뛰고 나머지로 진행)
        self.set_status("파일을 읽는 중...")
        all_tasks, all_boards, total = [], [], 0
        used, failed = [], []
        for p in self.paths:
            try:
                tasks, board = parse_report(p)
            except Exception as e:
                failed.append((os.path.basename(p), str(e)))
                continue
            if not tasks and not board:
                failed.append((os.path.basename(p), "업무 내용을 찾지 못했습니다."))
                continue
            all_tasks.append(tasks)
            all_boards.append(board)
            total += len(tasks)
            used.append(os.path.basename(p))

        if failed:
            msg = "아래 파일은 읽을 수 없어 제외됩니다:\n\n"
            for name, reason in failed:
                msg += f"• {name}\n   → {reason.splitlines()[0]}\n"
            if not used:
                msg += "\n읽을 수 있는 파일이 없어 중단합니다."
                messagebox.showerror("읽기 오류", msg, parent=self.root)
                self.set_status("읽을 수 있는 파일이 없습니다.", "#c0392b")
                return
            msg += f"\n나머지 {len(used)}개 파일로 계속 진행할까요?"
            if not messagebox.askyesno("일부 파일 제외", msg, parent=self.root):
                self.set_status("취소되었습니다.")
                return

        # 중복 확인 (체크 시): ① 데일리 업무와 겹침  ② 다른 구분 간 비슷한 업무
        if self.sim_var.get():
            # ① 데일리 업무와 겹치는 본문 업무 → 뺄지 확인
            self.set_status("데일리 업무 중복을 확인하는 중...")
            dcands = find_daily_duplicates(all_tasks)
            if dcands:
                drop = resolve_daily_duplicates(self.root, dcands)
                if drop is None:
                    self.set_status("취소되었습니다.")
                    return
                if drop:
                    all_tasks = [[t for t in tasks if id(t) not in drop] for tasks in all_tasks]

            # ② 다른 구분 간 비슷한 업무 → 합치기/개별 결정
            self.set_status("비슷한 업무를 확인하는 중...")
            pairs = find_similar_pairs(all_tasks)
            if pairs:
                drop = resolve_similar_pairs(self.root, pairs)
                if drop is None:
                    self.set_status("취소되었습니다.")
                    return
                if drop:
                    all_tasks = [[t for t in tasks if id(t) not in drop] for tasks in all_tasks]
            total = sum(len(ts) for ts in all_tasks)

        # 부서 공용 게시판 통합 → 파일마다 내용이 다르면 어느 것 반영할지 확인
        self.set_status("부서 공용 게시판을 통합하는 중...")
        merged_board = resolve_board(self.root, all_boards, used)
        if merged_board is None:
            self.set_status("취소되었습니다.")
            return

        # 업무 병합(같은 제목/입찰·조달) → 구분별 버킷
        buckets, extra_order = merge_tasks(all_tasks)

        # 진행률 100% 미만인데 금주 예정이 빈 업무 → 무엇을 넣을지 입력
        needs = find_missing_plans(buckets)
        if needs:
            self.set_status("금주 예정 업무 확인 중...")
            ok = resolve_missing_plans(self.root, needs)
            if ok is None:
                self.set_status("취소되었습니다.")
                return

        # 생성/저장
        self.set_status("보고서를 만드는 중...")
        wb = build_workbook(buckets, extra_order, merged_board, report_date)
        outdir = os.path.dirname(self.paths[0])
        outname = f"{부서명}_주간보고_{report_date.strftime('%Y%m%d')}.xlsx"
        outpath = os.path.join(outdir, outname)
        try:
            wb.save(outpath)
        except PermissionError:
            messagebox.showerror("저장 오류",
                f"결과 파일이 이미 열려 있어 저장할 수 없습니다.\n닫고 다시 시도하세요.\n\n{outpath}",
                parent=self.root)
            self.set_status("저장 실패(파일 열림).", "#c0392b")
            return

        # AI 요약본 (선택)
        ai_msg = ""
        ai_path = None
        if self.ai_var.get():
            key = load_api_key()
            if not key:
                key = self.set_api_key()
            if key:
                self.set_status("AI 요약을 만드는 중... (잠시 기다려 주세요)")
                try:
                    summary = generate_ai_summary(all_tasks, report_date, key)
                    swb = build_summary_workbook(summary, report_date)
                    ai_name = f"{부서명}_주간보고_AI요약_{report_date.strftime('%Y%m%d')}.xlsx"
                    ai_path = os.path.join(outdir, ai_name)
                    swb.save(ai_path)
                    ai_msg = f"\nAI 요약본: {ai_name}"
                except urllib.error.HTTPError as e:
                    apimsg = getattr(e, "_apimsg", "")
                    if not apimsg:
                        try:
                            apimsg = json.loads(e.read().decode("utf-8", "replace")).get("error", {}).get("message", "")
                        except Exception:
                            apimsg = ""
                    low = apimsg.lower()
                    if "credit balance" in low or "billing" in low:
                        hint = ("👉 Anthropic API 계정에 결제(크레딧)가 없습니다.\n"
                                "   console.anthropic.com → Plans & Billing 에서 크레딧을\n"
                                "   충전(또는 결제수단 등록)한 뒤 다시 시도하세요.\n"
                                "   ※ 이건 키 오류가 아니라 'API 사용 잔액 부족'입니다.")
                    elif e.code in (401, 403):
                        hint = "👉 API 키가 올바른지 확인하세요. (API 키 설정 버튼)"
                    elif "model" in low:
                        hint = f"👉 모델 문제일 수 있습니다. 현재 시도 모델: {', '.join(AI_MODELS)}"
                    else:
                        hint = ""
                    messagebox.showwarning("AI 요약 실패",
                        f"AI 요약본 생성 실패 (HTTP {e.code})\n"
                        f"{apimsg}\n\n{hint}\n\n기본 보고서는 정상 생성되었습니다.",
                        parent=self.root)
                except Exception as e:
                    messagebox.showwarning("AI 요약 실패",
                        f"AI 요약본 생성에 실패했습니다.\n{e}\n\n기본 보고서는 정상 생성되었습니다.",
                        parent=self.root)

        self.set_status(f"완료! 총 {total}건 수록 → {outname}", "#1f9d55")
        messagebox.showinfo("완료",
            f"보고서가 생성되었습니다.\n\n파일: {outname}\n수록 업무: 총 {total}건{ai_msg}\n저장 위치: {outdir}\n\n파일을 엽니다.",
            parent=self.root)
        try:
            os.startfile(outpath)
            if ai_path:
                os.startfile(ai_path)
        except Exception:
            pass

    def set_api_key(self):
        """API 키 입력받아 저장하고 반환. (이미 있으면 그 값을 보여줌)"""
        cur = load_api_key()
        key = simpledialog.askstring(
            "Claude API 키",
            "Claude API 키를 입력하세요 (sk-ant- 로 시작).\n"
            "한 번 입력하면 프로그램 옆에 저장되어 다음부터 자동 사용됩니다.\n"
            "키 발급: https://console.anthropic.com 에서 API Keys",
            initialvalue=cur, parent=self.root, show="*")
        if key and key.strip():
            save_api_key(key.strip())
            self.set_status("API 키를 저장했습니다.", "#1f3b57")
            return key.strip()
        return ""


def main():
    root = tk.Tk()
    ReportApp(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception:
        try:
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("오류", traceback.format_exc())
        except Exception:
            print(traceback.format_exc())
